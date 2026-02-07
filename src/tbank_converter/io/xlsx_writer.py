"""XLSX writer for double-entry bookkeeping operations."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from tbank_converter.domain.models import Report


# Column indices (1-based for openpyxl)
COL_DATE = 1          # A - Дата
COL_DEBIT = 2         # B - Дебет (Куда)
COL_CREDIT = 3        # C - Кредит (Откуда)
COL_AMOUNT = 4        # D - Сумма
COL_CURRENCY = 5      # E - Валюта
COL_DESCRIPTION = 6   # F - Описание
COL_COMMENT = 7       # G - Комментарий
COL_CATEGORY = 8      # H - Категория
COL_SUBCATEGORY = 9   # I - Подкатегория

# Column headers
HEADERS = [
    "Дата",
    "Дебет (Куда)",
    "Кредит (Откуда)",
    "Сумма",
    "Валюта",
    "Описание",
    "Комментарий",
    "Категория",
    "Подкатегория",
]

CURRENCY_SYMBOLS: dict[str, str] = {
    "RUB": "₽",
    "USD": "$",
    "EUR": "€",
    "GBP": "£",
    "CNY": "¥",
}


class XLSXWriter:
    """Writes Report to XLSX file with double-entry bookkeeping format."""

    def __init__(self, report: Report, currency_code: str = "RUB"):
        """Initialize writer.

        Args:
            report: Report object with operations.
            currency_code: ISO currency code (e.g. "RUB", "USD").
        """
        self.report = report
        self.currency_symbol = CURRENCY_SYMBOLS.get(currency_code, currency_code)

    def write(self, output_path: Path) -> None:
        """Write report to XLSX file.

        Args:
            output_path: Path for output XLSX file.
        """
        wb = Workbook()
        ws = wb.active

        # Set sheet name based on report period
        if self.report.period_start:
            sheet_name = self.report.period_start.strftime("%B%Y")
        else:
            sheet_name = "Транзакции"
        ws.title = sheet_name[:31]

        self._write_header(ws)
        self._write_data(ws)
        self._apply_styles(ws)
        self._set_column_widths(ws)

        wb.save(output_path)

    def _write_header(self, ws: Worksheet) -> None:
        """Write header row with column names."""
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _write_data(self, ws: Worksheet) -> None:
        """Write operation data rows."""
        for row_idx, op in enumerate(self.report.operations, start=2):
            # Date (formatted as yyyy-mm-dd)
            ws.cell(
                row=row_idx,
                column=COL_DATE,
                value=op.operation_date.strftime("%Y-%m-%d")
            )

            # Debit account
            ws.cell(row=row_idx, column=COL_DEBIT, value=op.debit_account)

            # Credit account
            ws.cell(row=row_idx, column=COL_CREDIT, value=op.credit_account)

            # Amount (always positive - absolute value)
            ws.cell(
                row=row_idx,
                column=COL_AMOUNT,
                value=float(abs(op.operation_amount))
            )

            # Currency
            ws.cell(row=row_idx, column=COL_CURRENCY, value=op.operation_currency)

            # Description
            ws.cell(row=row_idx, column=COL_DESCRIPTION, value=op.description)

            # Comment (empty for user to fill)
            ws.cell(row=row_idx, column=COL_COMMENT, value=op.comment)

            # Category (only for expenses)
            ws.cell(row=row_idx, column=COL_CATEGORY, value=op.category)

            # Subcategory (only for expenses)
            ws.cell(row=row_idx, column=COL_SUBCATEGORY, value=op.subcategory)

    def _apply_styles(self, ws: Worksheet) -> None:
        """Apply number formatting and alignment."""
        data_end_row = len(self.report.operations) + 1  # +1 for header

        # Amount column: number format with currency symbol
        number_format = f'#,##0.00 {self.currency_symbol}'
        for row in range(2, data_end_row + 1):
            cell = ws.cell(row=row, column=COL_AMOUNT)
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal="right")

    def _set_column_widths(self, ws: Worksheet) -> None:
        """Set column widths for better readability."""
        column_widths = {
            COL_DATE: 12,           # Дата
            COL_DEBIT: 20,          # Дебет (Куда)
            COL_CREDIT: 20,         # Кредит (Откуда)
            COL_AMOUNT: 15,         # Сумма
            COL_CURRENCY: 10,       # Валюта
            COL_DESCRIPTION: 40,    # Описание
            COL_COMMENT: 30,        # Комментарий
            COL_CATEGORY: 20,       # Категория
            COL_SUBCATEGORY: 18,    # Подкатегория
        }

        for col_idx, width in column_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
