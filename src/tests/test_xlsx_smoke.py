"""Smoke tests for XLSX generation."""

from datetime import datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from tbank_converter.domain.models import Operation, Report
from tbank_converter.io.xlsx_writer import XLSXWriter

from tests.helpers import make_operation


@pytest.fixture
def sample_report():
    """Create sample report with double-entry fields populated."""
    operations = [
        make_operation(
            operation_date=datetime(2026, 1, 15, 12, 0, 0),
            operation_amount=Decimal("-100.50"),
            payment_amount=Decimal("-100.50"),
            total_payment_amount=Decimal("-100.50"),
            bank_category="Супермаркеты",
            mcc="5411",
            description="Пятёрочка",
            bonus_count="0",
            debit_account="расходы",
            credit_account="Счёт ТБанка",
            category="продукты",
            subcategory="еда",
        ),
        make_operation(
            operation_date=datetime(2026, 1, 16, 14, 30, 0),
            operation_amount=Decimal("10000.00"),
            payment_amount=Decimal("10000.00"),
            total_payment_amount=Decimal("10000.00"),
            description="Зарплата",
            bonus_count="0",
            debit_account="Счёт ТБанка",
            credit_account="доходы",
        ),
    ]

    return Report(
        operations=operations,
        categories=["продукты"],
    )


def test_xlsx_creates_file(tmp_path, sample_report):
    """Test that XLSX file is created."""
    output_path = tmp_path / "test.xlsx"

    writer = XLSXWriter(sample_report)
    writer.write(output_path)

    assert output_path.exists()


def test_xlsx_has_correct_headers(tmp_path, sample_report):
    """Test that XLSX has correct column headers for double-entry format."""
    output_path = tmp_path / "test.xlsx"

    writer = XLSXWriter(sample_report)
    writer.write(output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    # Check 9 column headers
    assert ws["A1"].value == "Дата"
    assert ws["B1"].value == "Дебет (Куда)"
    assert ws["C1"].value == "Кредит (Откуда)"
    assert ws["D1"].value == "Сумма"
    assert ws["E1"].value == "Валюта"
    assert ws["F1"].value == "Описание"
    assert ws["G1"].value == "Комментарий"
    assert ws["H1"].value == "Категория"
    assert ws["I1"].value == "Подкатегория"

    wb.close()


def test_xlsx_data_values(tmp_path, sample_report):
    """Test that XLSX contains correct data values."""
    output_path = tmp_path / "test.xlsx"

    writer = XLSXWriter(sample_report)
    writer.write(output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    # Check first operation (expense)
    assert ws["A2"].value == "2026-01-15"  # Date
    assert ws["B2"].value == "расходы"  # Debit
    assert ws["C2"].value == "Счёт ТБанка"  # Credit
    assert ws["D2"].value == 100.50  # Amount (positive)
    assert ws["E2"].value == "RUB"  # Currency
    assert ws["F2"].value == "Пятёрочка"  # Description
    assert ws["G2"].value in ("", None)  # Comment (empty - openpyxl uses None for empty cells)
    assert ws["H2"].value == "продукты"  # Category
    assert ws["I2"].value == "еда"  # Subcategory

    # Check second operation (income)
    assert ws["A3"].value == "2026-01-16"  # Date
    assert ws["B3"].value == "Счёт ТБанка"  # Debit
    assert ws["C3"].value == "доходы"  # Credit
    assert ws["D3"].value == 10000.00  # Amount (positive)
    assert ws["H3"].value in ("", None)  # Category (empty for income)
    assert ws["I3"].value in ("", None)  # Subcategory (empty for income)

    wb.close()


def test_xlsx_sheet_name_from_date(tmp_path, sample_report):
    """Test that sheet name is set from report period."""
    output_path = tmp_path / "test.xlsx"

    writer = XLSXWriter(sample_report)
    writer.write(output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    # Sheet name should be formatted as "MonthYear" (e.g., "January2026")
    assert ws.title == "January2026"

    wb.close()


def test_xlsx_no_formulas(tmp_path, sample_report):
    """Test that XLSX has no formulas (summary table removed)."""
    output_path = tmp_path / "test.xlsx"

    writer = XLSXWriter(sample_report)
    writer.write(output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    # Check that there are no SUMIF/COUNTIF formulas anywhere
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                assert "SUMIF" not in cell.value.upper()
                assert "COUNTIF" not in cell.value.upper()

    wb.close()


def test_xlsx_amount_formatting(tmp_path, sample_report):
    """Test that amount column has correct number format."""
    output_path = tmp_path / "test.xlsx"

    writer = XLSXWriter(sample_report)
    writer.write(output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    # Check that D2 (amount) has pure number format
    amount_cell = ws["D2"]
    assert amount_cell.number_format == "#,##0.00"

    wb.close()
